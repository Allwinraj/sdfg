"""Generate rich, parser-friendly fixtures for product-feature.md journeys 1–3.

Excel/CSV sheets keep the header on row 1 so Architect ingest can detect schema.
Invoices.pdf is a machine extract (CSV lines) plus a visual register layout.
"""

from __future__ import annotations

import csv
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib.colors import HexColor, white
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

ROOT = Path(__file__).resolve().parent
NAVY = "1B365D"
GOLD = "C4A35A"
SLATE = "F4F1EA"
INK = "1C1917"
THIN = Border(
    left=Side(style="thin", color="D6D3D1"),
    right=Side(style="thin", color="D6D3D1"),
    top=Side(style="thin", color="D6D3D1"),
    bottom=Side(style="thin", color="D6D3D1"),
)


def money(n: Decimal | int | str | float) -> Decimal:
    return Decimal(str(n)).quantize(Decimal("0.01"))


def write_csv(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, lineterminator="\n")
        for row in rows:
            writer.writerow(row)


def write_xlsx(path: Path, sheet: str, headers: list[str], rows: list[list[object]], money_idx: list[int], date_idx: list[int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=NAVY)
    stripe = PatternFill("solid", fgColor=SLATE)
    body = Font(name="Calibri", size=10, color=INK)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN
    for r, row in enumerate(rows, 2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(r, c, value)
            cell.font = body
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            if r % 2 == 0:
                cell.fill = stripe
            if c - 1 in money_idx and isinstance(value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'
            if c - 1 in date_idx:
                cell.number_format = "YYYY-MM-DD"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    for i, header in enumerate(headers, 1):
        width = max(12, min(36, len(header) + 4))
        for r in range(2, min(len(rows) + 2, 12)):
            val = ws.cell(r, i).value
            if val is not None:
                width = max(width, min(42, len(str(val)) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width
    safe = "".join(ch if ch.isalnum() else "" for ch in sheet)[:18] or "Data"
    tab = Table(displayName=f"{safe}Tbl", ref=ws.dimensions)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    cover = wb.create_sheet("Pack notes", 1)
    cover["A1"] = "Helios Industrial — demo pack for Nexus Architect"
    cover["A1"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    cover["A3"] = "Working sheet is the first tab. Header row is row 1 so schema detection stays clean."
    cover.column_dimensions["A"].width = 92
    wb.save(path)


def draw_banner(canvas, doc, title: str, subtitle: str) -> None:
    canvas.saveState()
    canvas.setFillColor(HexColor("#" + NAVY))
    canvas.rect(0, letter[1] - 72, letter[0], 72, fill=1, stroke=0)
    canvas.setFillColor(HexColor("#" + GOLD))
    canvas.rect(0, letter[1] - 76, letter[0], 4, fill=1, stroke=0)
    canvas.setFillColor(white)
    canvas.setFont("Times-Bold", 16)
    canvas.drawString(54, letter[1] - 40, title)
    canvas.setFont("Times-Roman", 9)
    canvas.drawString(54, letter[1] - 56, subtitle)
    canvas.setFillColor(HexColor("#78716C"))
    canvas.setFont("Times-Roman", 8)
    canvas.drawString(54, 28, "Confidential — Helios Industrial demo dataset  ·  Not for production posting")
    canvas.restoreState()


def write_policy_pdf(path: Path, title: str, subtitle: str, sections: list[tuple[str, list[str]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    h = ParagraphStyle(
        "H",
        parent=styles["Heading2"],
        fontName="Times-Bold",
        textColor=HexColor("#" + NAVY),
        spaceBefore=14,
        spaceAfter=8,
    )
    body = ParagraphStyle(
        "B",
        parent=styles["BodyText"],
        fontName="Times-Roman",
        fontSize=10,
        leading=14,
        textColor=HexColor("#" + INK),
        spaceAfter=6,
    )
    doc = SimpleDocTemplate(
        str(path),
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=96,
        bottomMargin=48,
        title=title,
        author="Helios Industrial Controllership",
    )
    story: list = [
        Paragraph(title, ParagraphStyle("T", parent=styles["Title"], textColor=HexColor("#" + NAVY), fontName="Times-Bold")),
        Paragraph(subtitle, ParagraphStyle("S", parent=styles["Normal"], textColor=HexColor("#78716C"), fontName="Times-Italic")),
        Spacer(1, 12),
    ]
    for heading, paras in sections:
        story.append(Paragraph(heading, h))
        for para in paras:
            story.append(Paragraph(para, body))
    doc.build(
        story,
        onFirstPage=lambda c, d: draw_banner(c, d, "Helios Industrial", subtitle),
        onLaterPages=lambda c, d: draw_banner(c, d, "Helios Industrial", subtitle),
    )


def write_csv_extract_pdf(path: Path, title: str, rows: list[list[object]]) -> None:
    """CSV-only extractable text so /chat/upload can detect a table.

    Banner chrome is drawn as shapes only — title text would steal the first
    parsed lines and collapse the register into a single column.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    from reportlab.pdfgen.canvas import Canvas

    c = Canvas(str(path), pagesize=letter)
    c.setTitle(title)
    text_rows = [",".join(_csv_cell(v) for v in row) for row in rows]
    y = 0
    for i, line in enumerate(text_rows):
        if i == 0 or y < 40:
            if i:
                c.showPage()
            c.setFillColor(HexColor("#" + NAVY))
            c.rect(0, letter[1] - 28, letter[0], 28, fill=1, stroke=0)
            c.setFillColor(HexColor("#" + GOLD))
            c.rect(0, letter[1] - 32, letter[0], 4, fill=1, stroke=0)
            y = letter[1] - 48
            c.setFillColor(HexColor("#" + INK))
            c.setFont("Courier", 6.5)
        c.drawString(18, y, line[:172])
        y -= 9
    c.save()


def _csv_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    if isinstance(value, date):
        return value.isoformat()
    text = str(value)
    if any(ch in text for ch in ',|"\n'):
        return '"' + text.replace('"', '""') + '"'
    return text


def journey_invoice() -> None:
    out = ROOT / "journey-01-invoice-3way-matching"
    out.mkdir(parents=True, exist_ok=True)

    vendors = [
        ("V-1001", "Dell Inc", "Dell Computers", "Tier-1", "US12-338821"),
        ("V-1044", "SKF USA Inc", "SKF United States", "Tier-1", "US22-910044"),
        ("V-1102", "Bosch Rexroth Corp", "Robert Bosch Rexroth", "Tier-2", "US18-441102"),
        ("V-1188", "Fastenal Company", "Fastenal Co.", "Tier-1", "US41-094841"),
        ("V-1210", "Grainger Inc", "W.W. Grainger", "Tier-2", "US36-1150280"),
        ("V-1308", "Mitsubishi Electric US", "Mitsubishi Electric America", "Tier-1", "US95-330881"),
        ("V-1412", "Acme Tooling LLC", "Acme Tooling Inc", "Tier-3", "US87-220019"),
        ("V-1550", "Siemens Industry Inc", "Siemens Industrial", "Tier-1", "US13-2768159"),
    ]
    # Catalog of PO lines: planted exceptions called out in comments via scenario tags.
    catalog = [
        # perfect 3-way
        ("PO-24018", "10", 0, "SRV-7400", "PowerEdge R760 server", 4, "EA", "4820.00", "Net 45", "Plant-OH"),
        ("PO-24018", "20", 0, "WAR-3YR", "ProSupport 3-year", 4, "EA", "610.00", "Net 45", "Plant-OH"),
        ("PO-24022", "10", 1, "6205-2RS", "Deep groove ball bearing", 240, "EA", "18.40", "Net 30", "Plant-IN"),
        ("PO-24022", "20", 1, "6310-2Z", "Shielded bearing 50mm", 80, "EA", "42.15", "Net 30", "Plant-IN"),
        ("PO-24031", "10", 2, "4WE6", "Directional valve CETOP 3", 12, "EA", "318.50", "Net 30", "Plant-OH"),
        ("PO-24031", "20", 2, "A10VSO", "Axial piston pump 28cc", 2, "EA", "2140.00", "Net 30", "Plant-OH"),
        ("PO-24040", "10", 3, "SHCS-M8", "Socket cap screw M8x30", 5000, "EA", "0.18", "Net 15", "Plant-TX"),
        ("PO-24040", "20", 3, "NUT-M8", "Locknut M8 zinc", 5000, "EA", "0.07", "Net 15", "Plant-TX"),
        ("PO-24044", "10", 4, "PPE-19", "Arc-flash kit CAT 2", 18, "EA", "214.00", "Net 30", "Plant-TX"),
        ("PO-24051", "10", 5, "FR-A840", "VFD 15kW 480V", 6, "EA", "1688.00", "Net 45", "Plant-OH"),
        ("PO-24051", "20", 5, "FR-CAB", "VFD EMC cable 15m", 6, "EA", "96.40", "Net 45", "Plant-OH"),
        ("PO-24058", "10", 7, "6ES7-315", "S7-300 CPU 315-2 PN/DP", 3, "EA", "2895.00", "Net 30", "Plant-IN"),
        ("PO-24058", "20", 7, "6ES7-322", "DO 16x24V module", 8, "EA", "412.00", "Net 30", "Plant-IN"),
        ("PO-24062", "10", 6, "JIG-441", "Custom weld fixture plate", 1, "EA", "1880.00", "Net 30", "Plant-OH"),
        ("PO-24067", "10", 0, "MON-27", "27in UHD production monitor", 10, "EA", "428.00", "Net 45", "Plant-TX"),
        ("PO-24070", "10", 1, "SNL-520", "Plummer block housing", 16, "EA", "126.80", "Net 30", "Plant-IN"),
        ("PO-24073", "10", 3, "TAP-M10", "Spiral flute tap M10", 40, "EA", "14.25", "Net 15", "Plant-TX"),
        ("PO-24077", "10", 4, "GLV-NIT", "Nitrile gloves XL 100ct", 60, "BX", "22.40", "Net 30", "Plant-OH"),
        ("PO-24081", "10", 5, "GT-200", "GOT2000 HMI 10in", 4, "EA", "980.00", "Net 45", "Plant-OH"),
        ("PO-24088", "10", 7, "3RT2", "Contactor 32A", 24, "EA", "68.90", "Net 30", "Plant-IN"),
        ("PO-24091", "10", 2, "HOSE-2S", "2-wire hydraulic hose 10m", 20, "EA", "54.00", "Net 30", "Plant-TX"),
        ("PO-24095", "10", 1, "OIL-32", "Hydraulic ISO 32 55gal", 8, "DR", "312.00", "Net 30", "Plant-OH"),
        ("PO-24102", "10", 0, "NB-5G", "5G industrial router", 2, "EA", "1240.00", "Net 45", "Plant-OH"),
        ("PO-24108", "10", 4, "EAR-NBR", "Disposable earplugs 200ct", 40, "BX", "9.85", "Net 30", "Plant-TX"),
        ("PO-24114", "10", 3, "BIT-1/2", "Cobalt drill 1/2in", 80, "EA", "6.40", "Net 15", "Plant-IN"),
        ("PO-24120", "10", 7, "ET200SP", "IM 155-6 PN ST", 5, "EA", "534.00", "Net 30", "Plant-OH"),
        ("PO-24125", "10", 5, "SC-400", "Servo motor 400W", 8, "EA", "410.00", "Net 45", "Plant-IN"),
        ("PO-24130", "10", 2, "FIL-10", "Return line filter 10µm", 12, "EA", "86.25", "Net 30", "Plant-OH"),
        ("PO-24136", "10", 6, "DIE-09", "Progressive die insert", 2, "EA", "760.00", "Net 30", "Plant-OH"),
        ("PO-24142", "10", 1, "SEAL-V", "Viton shaft seal 45mm", 100, "EA", "4.18", "Net 30", "Plant-TX"),
    ]

    po_date = date(2024, 4, 8)
    po_headers = [
        "po_number",
        "po_line_id",
        "vendor_id",
        "vendor_name",
        "vendor_tier",
        "item_sku",
        "item_description",
        "qty_ordered",
        "uom",
        "unit_price",
        "line_amount",
        "currency",
        "po_date",
        "buyer",
        "plant",
        "gl_account",
        "cost_center",
        "payment_terms",
        "incoterms",
        "status",
    ]
    po_rows: list[list[object]] = []
    for i, rec in enumerate(catalog):
        po, line, vi, sku, desc, qty, uom, price, terms, plant = rec
        v = vendors[vi]
        unit = money(price)
        qty_d = Decimal(qty)
        po_rows.append(
            [
                po,
                line,
                v[0],
                v[1],
                v[3],
                sku,
                desc,
                int(qty_d),
                uom,
                unit,
                money(unit * qty_d),
                "USD",
                po_date + timedelta(days=i % 12),
                ["J. Patel", "M. Chen", "A. Rossi", "S. Okonkwo"][i % 4],
                plant,
                "2110-6400",
                "CC-OPS-14",
                terms,
                "DAP",
                "Open",
            ]
        )

    # Extra clean lines to thicken the extract
    extras = [
        ("PO-24150", "10", 3, "WASHER-M8", "Flat washer M8", 8000, "EA", "0.03", "Net 15", "Plant-TX"),
        ("PO-24150", "20", 3, "LOCT-243", "Threadlocker 243 50ml", 24, "EA", "11.80", "Net 15", "Plant-TX"),
        ("PO-24161", "10", 4, "BOOTS-11", "Composite-toe boot 11", 22, "PR", "128.00", "Net 30", "Plant-OH"),
        ("PO-24168", "10", 7, "SCALANCE", "Managed switch 8-port", 3, "EA", "890.00", "Net 30", "Plant-IN"),
        ("PO-24174", "10", 0, "DOCK-USBC", "USB-C docking station", 14, "EA", "219.00", "Net 45", "Plant-TX"),
        ("PO-24180", "10", 1, "GREASE-2", "Lithium grease 14oz", 48, "EA", "8.90", "Net 30", "Plant-OH"),
        ("PO-24186", "10", 5, "MR-J4", "Servo amplifier 400W", 8, "EA", "620.00", "Net 45", "Plant-IN"),
        ("PO-24192", "10", 2, "ACC-10", "Hydraulic accumulator 1L", 4, "EA", "355.00", "Net 30", "Plant-OH"),
    ]
    base_len = len(po_rows)
    for j, rec in enumerate(extras):
        i = base_len + j
        po, line, vi, sku, desc, qty, uom, price, terms, plant = rec
        v = vendors[vi]
        unit = money(price)
        qty_d = Decimal(qty)
        po_rows.append(
            [
                po,
                line,
                v[0],
                v[1],
                v[3],
                sku,
                desc,
                int(qty_d),
                uom,
                unit,
                money(unit * qty_d),
                "USD",
                po_date + timedelta(days=12 + j),
                ["J. Patel", "M. Chen"][j % 2],
                plant,
                "2110-6400",
                "CC-OPS-14",
                terms,
                "DAP",
                "Open",
            ]
        )

    write_xlsx(
        out / "PO_Export.xlsx",
        "PO_Export",
        po_headers,
        po_rows,
        money_idx=[9, 10],
        date_idx=[12],
    )

    gr_headers = [
        "gr_number",
        "gr_date",
        "po_number",
        "po_line_id",
        "vendor_name",
        "item_sku",
        "qty_received",
        "uom",
        "unit_price",
        "received_amount",
        "warehouse",
        "packing_slip",
        "receiver",
        "inspection_status",
    ]
    skip_gr = {("PO-24062", "10"), ("PO-24136", "10")}  # unreceived — decision path
    partial = {("PO-24022", "10"): 180}  # ordered 240
    gr_rows: list[list[object]] = []
    gn = 88020
    for rec in po_rows:
        key = (str(rec[0]), str(rec[1]))
        if key in skip_gr:
            continue
        qty = partial.get(key, int(rec[7]))
        unit = money(rec[9])
        vi = next(v for v in vendors if v[0] == rec[2])
        vendor_on_gr = vi[2] if rec[0] in {"PO-24018", "PO-24051", "PO-24102"} else vi[1]
        gn += 1
        day = rec[12] if isinstance(rec[12], date) else date(2024, 4, 20)
        gr_rows.append(
            [
                f"GR-{gn}",
                day + timedelta(days=6),
                rec[0],
                rec[1],
                vendor_on_gr,
                rec[5],
                qty,
                rec[8],
                unit,
                money(unit * Decimal(qty)),
                rec[14] + "-WH1",
                f"PS-{gn}",
                ["L. Nguyen", "K. Brooks", "P. Singh"][gn % 3],
                "Accepted",
            ]
        )
    write_csv(out / "Goods_Receipts.csv", [gr_headers, *gr_rows])

    inv_headers = [
        "invoice_number",
        "invoice_date",
        "po_number",
        "po_line_id",
        "vendor_name",
        "vendor_tax_id",
        "item_sku",
        "qty_invoiced",
        "uom",
        "unit_price",
        "invoice_amount",
        "tax_amount",
        "total_amount",
        "currency",
        "payment_terms",
        "due_date",
        "ap_clerk",
    ]
    # Amount exceptions (matcher keeps $0.02 as near-match evidence; math applies 2% / $50)
    amount_override = {
        ("PO-24018", "10"): money("19280.02"),  # 4 * 4820 = 19280.00 → +0.02
        ("PO-24031", "20"): money("4480.00"),  # 2 * 2140 = 4280 → ~4.7% over (flag)
        ("PO-24058", "10"): money("8685.00"),  # 3 * 2895 = 8685 exact
    }
    skip_inv = {("PO-24142", "10")}  # received, not yet billed
    inv_rows: list[list[object]] = []
    seq = 14000
    for rec in po_rows:
        key = (str(rec[0]), str(rec[1]))
        if key in skip_inv:
            continue
        vi = next(v for v in vendors if v[0] == rec[2])
        vendor_on_inv = vi[2]  # semantic variants vs PO legal name
        qty = int(rec[7])
        unit = money(rec[9])
        default_amt = money(unit * Decimal(qty))
        amt = amount_override.get(key, default_amt)
        tax = money(amt * Decimal("0.00"))  # tax-exempt industrial, keeps schema
        seq += 1
        inv_date = (rec[12] if isinstance(rec[12], date) else date(2024, 4, 22)) + timedelta(days=11)
        inv_rows.append(
            [
                f"INV-24-{seq}",
                inv_date,
                rec[0],
                rec[1],
                vendor_on_inv,
                vi[4],
                rec[5],
                qty,
                rec[8],
                unit,
                amt,
                tax,
                money(amt + tax),
                "USD",
                rec[17],
                inv_date + timedelta(days=30),
                ["R. Alvarez", "T. Berg", "H. Cole"][seq % 3],
            ]
        )
    # Unreceived Acme die (PO-24136) invoiced $760 — over $500, not Tier-1 → buyer review
    # Unreceived fixture (PO-24062) invoiced $420 — under $500, but Tier-3 (not Tier-1)
    # Extra small Fastenal invoice without GR for Tier-1 auto-approve story
    inv_rows.append(
        [
            "INV-24-14901",
            date(2024, 5, 3),
            "PO-24999",
            "10",
            "Fastenal Co.",
            vendors[3][4],
            "MISC-BIN",
            1,
            "EA",
            money("186.00"),
            money("186.00"),
            money("0.00"),
            money("186.00"),
            "USD",
            "Net 15",
            date(2024, 5, 18),
            "R. Alvarez",
        ]
    )

    write_xlsx(out / "Invoices.xlsx", "Invoices", inv_headers, inv_rows, money_idx=[9, 10, 11, 12], date_idx=[1, 15])
    write_csv_extract_pdf(out / "Invoices.pdf", "Helios AP — vendor invoice register (May 2024)", [inv_headers, *inv_rows])

    write_policy_pdf(
        out / "AP_Procurement_Policy_2024.pdf",
        "AP & Procurement Policy 2024",
        "Document AP-SOP-2024.07  ·  Effective 1 January 2024  ·  Mid-market manufacturing",
        [
            (
                "1. Purpose and domain",
                [
                    "This policy governs three-way match of purchase orders, goods receipts, and vendor invoices for Helios Industrial (P2P). All amounts are United States dollars. Multi-currency conversion is not authorized in this revision.",
                    "Match keys for structural alignment are <b>PO Number</b> and <b>PO Line Item ID</b>. Vendor legal names may differ across ERP, packing slips, and invoices (for example Dell Inc versus Dell Computers). Semantic normalization is required; do not fail a match solely on vendor spelling.",
                ],
            ),
            (
                "2. Price and quantity tolerance (Math Engine — not Matcher)",
                [
                    "The Matcher records evidence only. Numeric gates are applied after match: allow the lesser of <b>2% of PO line amount</b> or <b>USD 50.00</b>. Formula: MIN(0.02 × po_line_amount, 50.00). Variances inside the gate stay on the matched path; larger gaps flag for buyer review.",
                    "A two-cent rounding difference is a near-match, not an automatic reject.",
                ],
            ),
            (
                "3. Unreceived invoices (Decision / approval gateway)",
                [
                    "If an invoice has no matching goods receipt and the invoice total is under USD 500.00 <b>and</b> the vendor is classified Tier-1 in the vendor master, Accounts Payable may auto-approve.",
                    "All other unreceived invoices route to buyer review, including Tier-2/Tier-3 vendors and any invoice at or above USD 500.00.",
                ],
            ),
            (
                "4. Vendor tier snapshot (knowledge)",
                [
                    "Tier-1: Dell Inc, SKF USA Inc, Fastenal Company, Mitsubishi Electric US, Siemens Industry Inc.",
                    "Tier-2: Bosch Rexroth Corp, Grainger Inc. Tier-3: Acme Tooling LLC (custom tooling — no auto-approve on unreceived invoices).",
                ],
            ),
            (
                "5. Output",
                [
                    "Produce a multi-tab Excel workbook: Summary, Matched, Exceptions. Slack or email dispatch is outside this operating procedure for the current control environment.",
                ],
            ),
        ],
    )


def _add_bus_days(start: date, days: int) -> date:
    step = 1 if days >= 0 else -1
    left = abs(days)
    cur = start
    while left:
        cur += timedelta(days=step)
        if cur.weekday() < 5:
            left -= 1
    return cur


def journey_bank() -> None:
    out = ROOT / "journey-02-bank-reconciliation"
    out.mkdir(parents=True, exist_ok=True)
    opening = money("2450188.42")
    # Bank statement May 2024 — match on amount + reference; date may drift ±2 business days.
    bank_headers = [
        "statement_date",
        "value_date",
        "bank_txn_id",
        "txn_type",
        "debit_credit",
        "amount",
        "stated_running_balance",
        "reference_number",
        "description",
        "check_number",
        "counterparty",
        "branch",
        "currency",
    ]
    gl_headers = [
        "posting_date",
        "document_number",
        "gl_account",
        "gl_account_name",
        "amount",
        "debit_credit",
        "reference_number",
        "source_system",
        "customer_or_vendor",
        "cost_center",
        "cleared_flag",
        "value_date",
        "currency",
        "narration",
    ]

    events: list[dict] = []
    # Opening memo
    events.append({"kind": "open", "d": date(2024, 5, 1), "amt": Decimal("0"), "ref": "OPEN-MAY", "desc": "Brought forward"})

    # Daily store deposits (1:1)
    for i, amt in enumerate(
        [
            "128440.15",
            "131002.80",
            "119880.40",
            "142110.22",
            "138904.11",
            "125550.00",
            "133280.75",
            "140002.18",
            "129440.90",
            "136770.05",
            "121008.40",
            "144220.18",
            "130441.00",
            "127880.55",
            "139104.20",
            "118760.33",
            "141908.77",
            "132440.10",
            "126005.80",
            "135770.42",
            "122980.00",
            "148110.65",
        ]
    ):
        events.append(
            {
                "kind": "dep",
                "d": date(2024, 5, 1) + timedelta(days=i + 1),
                "amt": money(amt),
                "ref": f"DEP-MAY-{1001 + i}",
                "desc": "Card settlement — Helios Retail POS",
                "cp": "Helios Stores NA",
            }
        )

    # Allocation: one bank deposit covers three AR receipts
    events.append(
        {
            "kind": "alloc_bank",
            "d": date(2024, 5, 14),
            "amt": money("48500.00"),
            "ref": "WIRE-NEX-5510",
            "desc": "Incoming wire — customer remittance batch",
            "cp": "Northwind Wholesale",
        }
    )
    events.append({"kind": "alloc_gl", "d": date(2024, 5, 13), "amt": money("20000.00"), "ref": "WIRE-NEX-5510", "doc": "AR-88421", "cp": "Northwind Wholesale"})
    events.append({"kind": "alloc_gl", "d": date(2024, 5, 13), "amt": money("18500.00"), "ref": "WIRE-NEX-5510", "doc": "AR-88422", "cp": "Northwind Wholesale"})
    events.append({"kind": "alloc_gl", "d": date(2024, 5, 15), "amt": money("10000.00"), "ref": "WIRE-NEX-5510", "doc": "AR-88423", "cp": "Northwind Wholesale"})

    # Payroll, vendors, ACH
    for i, (amt, ref, desc, cp, t) in enumerate(
        [
            ("882140.00", "ACH-PAY-0515", "Bi-weekly payroll funding", "ADP Wage Pay", "wd"),
            ("44120.66", "ACH-UTIL-88", "Electric — plants OH/IN/TX", "Midwest Power", "wd"),
            ("12880.00", "ACH-LEASE-04", "Distribution center lease", "Prologis REIT", "wd"),
            ("6720.45", "FEE-ANA-MAY", "Account analysis fee", "First National Treasury", "wd"),
            ("25000.00", "CHK-441900", "Refund — overpaid freight", "XPO Logistics", "dep"),
            ("9104.22", "ACH-INS-Q2", "Property insurance Q2", "Chubb North America", "wd"),
            ("33450.00", "WIRE-TAX-941", "Federal 941 remittance", "US Treasury", "wd"),
            ("15600.80", "ACH-BENE-HSA", "HSA employer funding", "Fidelity Benefits", "wd"),
        ]
    ):
        events.append(
            {
                "kind": t,
                "d": date(2024, 5, 6) + timedelta(days=i * 2),
                "amt": money(amt),
                "ref": ref,
                "desc": desc,
                "cp": cp,
            }
        )

    # Stale checks issued April (uncleared > 30 days as of 31 May)
    events.append(
        {
            "kind": "stale_gl",
            "d": date(2024, 4, 12),
            "amt": money("8750.00"),
            "ref": "CHK-440112",
            "desc": "Outstanding check — tooling supplier",
            "cp": "Acme Tooling LLC",
            "check": "440112",
        }
    )
    events.append(
        {
            "kind": "stale_gl",
            "d": date(2024, 4, 18),
            "amt": money("4120.55"),
            "ref": "CHK-440188",
            "desc": "Outstanding check — freight claim",
            "cp": "Old Dominion Freight",
            "check": "440188",
        }
    )

    # 1:1 operating items with ±2 business day drift on GL
    twins = [
        (date(2024, 5, 3), "16880.40", "LOCKBOX-2291", "Lockbox — Midwest dealers", "Dealer AR"),
        (date(2024, 5, 7), "22104.00", "LOCKBOX-2298", "Lockbox — East dealers", "Dealer AR"),
        (date(2024, 5, 9), "990.00", "NSF-2291", "Returned item NSF", "Dealer AR"),
        (date(2024, 5, 16), "54000.00", "WIRE-EXPORT-9", "Export collection", "Helios Mexico SA"),
        (date(2024, 5, 20), "3122.18", "ACH-COURIER", "Overnight freight", "FedEx Corporate"),
        (date(2024, 5, 21), "17840.00", "ZBA-SWEEP", "ZBA residual sweep in", "First National"),
        (date(2024, 5, 22), "6400.00", "ACH-JANITOR", "Facilities services", "ABM Industry"),
        (date(2024, 5, 23), "21004.75", "LOCKBOX-2310", "Lockbox — West dealers", "Dealer AR"),
        (date(2024, 5, 28), "4550.00", "WIRE-REBATE", "Vendor rebate — Fastenal", "Fastenal Company"),
        (date(2024, 5, 29), "11880.00", "ACH-CLOUD", "ERP cloud subscription", "SAP America"),
        (date(2024, 5, 30), "775.40", "FEE-WIRE-OUT", "Wire origination fee", "First National Treasury"),
        (date(2024, 5, 31), "250000.00", "SWEEP-MMF", "Overnight money-market sweep", "Federated Hermes"),
    ]
    for i, (d, amt, ref, desc, cp) in enumerate(twins):
        typ = "wd" if ref.startswith("ACH") or ref.startswith("FEE") or ref.startswith("NSF") else "dep"
        if ref.startswith("NSF") or "FEE" in ref or ref.startswith("ACH-COURIER") or ref.startswith("ACH-JANITOR") or ref.startswith("ACH-CLOUD"):
            typ = "wd"
        if ref.startswith("SWEEP-MMF"):
            typ = "wd"
        events.append({"kind": typ, "d": d, "amt": money(amt), "ref": ref, "desc": desc, "cp": cp, "drift": (i % 5) - 2})

    bank_rows: list[list[object]] = []
    gl_rows: list[list[object]] = []
    balance = opening
    bid = 70000
    gid = 90000
    bank_rows.append(
        [
            date(2024, 5, 1),
            date(2024, 5, 1),
            "BNK-70000",
            "opening",
            "CR",
            opening,
            opening,
            "OPEN-MAY",
            "Opening collected balance",
            "",
            "First National — 4481",
            "Cincinnati OH",
            "USD",
        ]
    )
    gl_rows.append(
        [
            date(2024, 5, 1),
            "GL-90000",
            "111000",
            "Cash — operating",
            opening,
            "DR",
            "OPEN-MAY",
            "S/4HANA",
            "Internal",
            "CC-TREAS",
            "Y",
            date(2024, 5, 1),
            "USD",
            "Balance brought forward",
        ]
    )

    for ev in events:
        if ev["kind"] in {"open", "alloc_gl", "stale_gl"}:
            continue
        bid += 1
        d = ev["d"]
        amt = ev["amt"]
        dc = "CR" if ev["kind"] in {"dep", "alloc_bank"} else "DR"
        if dc == "CR":
            balance = money(balance + amt)
        else:
            balance = money(balance - amt)
        check = ev.get("check", "")
        if str(ev.get("ref", "")).startswith("CHK-"):
            check = str(ev["ref"]).replace("CHK-", "")
        ttype = {"dep": "deposit", "wd": "withdrawal", "alloc_bank": "deposit"}.get(ev["kind"], ev["kind"])
        bank_rows.append(
            [
                d,
                d,
                f"BNK-{bid}",
                ttype,
                dc,
                amt,
                balance,
                ev["ref"],
                ev["desc"],
                check,
                ev.get("cp", ""),
                "Cincinnati OH",
                "USD",
            ]
        )
        if ev["kind"] == "alloc_bank":
            continue
        gid += 1
        drift = int(ev.get("drift", 0))
        gl_day = _add_bus_days(d, drift)
        gdc = "DR" if dc == "CR" else "CR"
        gl_rows.append(
            [
                gl_day,
                f"GL-{gid}",
                "111000",
                "Cash — operating",
                amt,
                gdc,
                ev["ref"],
                "S/4HANA",
                ev.get("cp", ""),
                "CC-TREAS",
                "Y",
                d,
                "USD",
                ev["desc"],
            ]
        )

    for ev in events:
        if ev["kind"] != "alloc_gl":
            continue
        gid += 1
        gl_rows.append(
            [
                ev["d"],
                ev["doc"],
                "111000",
                "Cash — operating",
                ev["amt"],
                "DR",
                ev["ref"],
                "S/4HANA",
                ev.get("cp", ""),
                "CC-AR",
                "Y",
                ev["d"],
                "USD",
                "Split cash application against open invoices",
            ]
        )

    for ev in events:
        if ev["kind"] != "stale_gl":
            continue
        gid += 1
        gl_rows.append(
            [
                ev["d"],
                f"GL-{gid}",
                "111000",
                "Cash — operating",
                ev["amt"],
                "CR",
                ev["ref"],
                "S/4HANA",
                ev.get("cp", ""),
                "CC-TREAS",
                "N",
                ev["d"],
                "USD",
                ev["desc"] + " — still outstanding; flag if older than 30 days per SOP",
            ]
        )

    # Bank-only residual (unapplied analysis remainder already in FEE); GL-only book cash in transit
    gid += 1
    gl_rows.append(
        [
            date(2024, 5, 31),
            f"GL-{gid}",
            "112100",
            "Cash in transit",
            money("2400.00"),
            "DR",
            "CIT-MAY-31",
            "S/4HANA",
            "Stores — late bag",
            "CC-TREAS",
            "N",
            date(2024, 5, 31),
            "USD",
            "Store deposit recorded 31 May, not on bank statement until June 3",
        ]
    )

    write_csv(out / "Bank_Statement_May.csv", [bank_headers, *bank_rows])
    write_xlsx(
        out / "GL_Cash_Export.xlsx",
        "GL_Cash",
        gl_headers,
        gl_rows,
        money_idx=[4],
        date_idx=[0, 11],
    )
    write_policy_pdf(
        out / "Treasury_Recon_SOP.pdf",
        "Treasury Bank Reconciliation SOP",
        "Document TR-SOP-441  ·  Retail R2R close  ·  First National operating account ****4481",
        [
            (
                "1. Match rule",
                [
                    "Match bank statement lines to general-ledger cash postings on <b>Amount</b> and <b>Reference Number</b>. Posting date may differ from statement date by up to <b>two business days</b> (temporal window). Single-currency USD only.",
                    "Customer remittances may land as one bank deposit covering several AR documents. Use allocation (M:N) matching and emit unapplied cash on the residuals port.",
                ],
            ),
            (
                "2. Running balance",
                [
                    "Computed collected balance = previous stated balance + deposits − withdrawals, row by row in statement order. Flag any row where computed balance differs from the bank’s stated running balance.",
                ],
            ),
            (
                "3. Stale items",
                [
                    "Uncleared checks and deposits in transit older than <b>30 calendar days</b> are high risk. Cite this SOP in the decision explanation. Outstanding April checks CHK-440112 (USD 8,750.00) and CHK-440188 (USD 4,120.55) are in-scope examples for the May close.",
                ],
            ),
            (
                "4. Output",
                [
                    "Certified PDF audit report with match-status summary and a controller sign-off block. Live chat or ticket dispatch is not part of this SOP.",
                ],
            ),
        ],
    )


def journey_intercompany() -> None:
    out = ROOT / "journey-03-intercompany-ar-ap"
    out.mkdir(parents=True, exist_ok=True)
    entities = [
        ("US-HOLD", "Helios Holdings Inc", "Helios Holdings", "US", "S/4"),
        ("US-MFG", "Helios Manufacturing USA", "Helios Mfg USA", "US", "S/4"),
        ("DE-NEX", "Nexus Germany GmbH", "DE-Nexus Corp", "DE", "SAP ECC"),
        ("UK-LTD", "Helios UK Ltd", "Helios United Kingdom", "UK", "S/4"),
        ("FR-SAS", "Helios France SAS", "Helios FR SAS", "FR", "S/4"),
        ("NL-BV", "Helios Netherlands BV", "Helios NL BV", "NL", "S/4"),
        ("SG-PTE", "Helios Singapore Pte Ltd", "Helios SG Pte", "SG", "S/4"),
        ("IN-PVT", "Helios India Pvt Ltd", "Helios India Private", "IN", "S/4"),
        ("CN-WFOE", "Helios China Co Ltd", "Helios China WFOE", "CN", "Kingdee"),
        ("MX-SA", "Helios Mexico SA de CV", "Helios Mexico SA", "MX", "S/4"),
        ("BR-LTDA", "Helios Brasil Ltda", "Helios Brazil Ltda", "BR", "TOTVS"),
        ("AU-PTY", "Helios Australia Pty Ltd", "Helios AU Pty", "AU", "S/4"),
    ]
    headers = [
        "entity_code",
        "entity_legal_name",
        "local_erp_name",
        "counterparty_code",
        "counterparty_name_local",
        "direction",
        "document_no",
        "document_date",
        "posting_date",
        "amount",
        "currency",
        "invoice_ref",
        "narrative",
        "local_doc_id",
        "trading_partner",
        "ic_account",
    ]

    def name_for(code: str, viewer: str) -> str:
        rec = next(e for e in entities if e[0] == code)
        if viewer == "DE-NEX" and code == "DE-NEX":
            return rec[1]
        if code == "DE-NEX" and viewer != "DE-NEX":
            return "DE-Nexus Corp"
        if viewer == "CN-WFOE" and code == "US-HOLD":
            return "Helios Holdings USA"
        return rec[1] if viewer in {code, "US-HOLD"} else rec[2]

    # Paired flows: AR on A must match AP on B. Amounts sized toward ~$18.4M volume.
    pairs: list[tuple] = []

    def pair(a, b, amt, doc, d1, d2, note, ar_inv, ap_inv=None):
        pairs.append((a, b, money(amt), doc, d1, d2, note, ar_inv, ap_inv or ar_inv))

    pair("US-MFG", "DE-NEX", "1850000.00", "IC-2401", date(2024, 1, 12), date(2024, 1, 12), "Finished goods — servo frames", "FG-DE-441")
    pair("DE-NEX", "US-MFG", "640000.00", "IC-2402", date(2024, 1, 18), date(2024, 1, 19), "Engineering recharge", "ENG-US-18")
    pair("US-HOLD", "UK-LTD", "2200000.00", "IC-2403", date(2024, 1, 20), date(2024, 1, 20), "Management fee Q1", "MGMT-Q1-UK")
    pair("UK-LTD", "FR-SAS", "410000.00", "IC-2404", date(2024, 1, 22), date(2024, 1, 22), "Shared service — EMEA AR hub", "SSC-FR-09")
    pair("FR-SAS", "NL-BV", "288400.00", "IC-2405", date(2024, 1, 24), date(2024, 1, 24), "Rotterdam DC fulfillment", "DC-NL-224")
    pair("NL-BV", "DE-NEX", "915000.00", "IC-2406", date(2024, 1, 28), date(2024, 1, 28), "EU spare-parts pool", "SP-DE-77")
    pair("SG-PTE", "IN-PVT", "760000.00", "IC-2407", date(2024, 1, 29), date(2024, 1, 29), "APAC tooling", "TL-IN-12")
    pair("IN-PVT", "CN-WFOE", "1330000.00", "IC-2408", date(2024, 1, 30), date(2024, 1, 30), "Contract manufacturing", "CM-CN-80")
    pair("CN-WFOE", "AU-PTY", "540000.00", "IC-2409", date(2024, 1, 30), date(2024, 1, 31), "ANZ finished goods", "FG-AU-03")
    pair("MX-SA", "US-MFG", "980000.00", "IC-2410", date(2024, 1, 15), date(2024, 1, 15), "Maquila assemblies", "MX-US-66")
    pair("BR-LTDA", "US-HOLD", "420000.00", "IC-2411", date(2024, 1, 16), date(2024, 1, 16), "Royalty — Latin America", "RY-BR-1")
    pair("AU-PTY", "SG-PTE", "310000.00", "IC-2412", date(2024, 1, 17), date(2024, 1, 17), "Regional HQ allocation", "HQ-SG-4")
    pair("US-MFG", "MX-SA", "175000.00", "IC-2413", date(2024, 1, 19), date(2024, 1, 19), "Steel coil recharge", "RM-MX-2")
    pair("UK-LTD", "US-HOLD", "860000.00", "IC-2414", date(2024, 1, 21), date(2024, 1, 21), "Dividend clearing", "DIV-UK-24")
    pair("DE-NEX", "NL-BV", "205000.00", "IC-2415", date(2024, 1, 25), date(2024, 1, 25), "IP licence Q1", "IP-NL-1")
    pair("FR-SAS", "UK-LTD", "99000.00", "IC-2416", date(2024, 1, 26), date(2024, 1, 26), "Paris demo center", "DM-UK-8")
    pair("IN-PVT", "SG-PTE", "148000.00", "IC-2417", date(2024, 1, 27), date(2024, 1, 27), "Shared IT", "IT-SG-5")
    pair("CN-WFOE", "US-MFG", "2100000.00", "IC-2418", date(2024, 1, 28), date(2024, 1, 28), "Power electronics", "PE-US-19")
    pair("BR-LTDA", "MX-SA", "67000.00", "IC-2419", date(2024, 1, 29), date(2024, 1, 29), "Spanish documentation pack", "DOC-MX")
    pair("AU-PTY", "US-HOLD", "255000.00", "IC-2420", date(2024, 1, 11), date(2024, 1, 11), "Insurance recharge", "INS-AU")

    extras = [
        ("US-HOLD", "US-MFG", "88000.00", "Shared services — payroll"),
        ("US-MFG", "UK-LTD", "52000.00", "Export packing surcharge"),
        ("DE-NEX", "FR-SAS", "31000.00", "Warranty reserve transfer"),
        ("UK-LTD", "NL-BV", "27500.00", "North Sea spares"),
        ("FR-SAS", "DE-NEX", "19400.00", "Alsace tooling hours"),
        ("NL-BV", "UK-LTD", "16200.00", "Brokering commission"),
        ("SG-PTE", "AU-PTY", "44800.00", "APAC marketing fund"),
        ("IN-PVT", "US-MFG", "67300.00", "Noida engineering hours"),
        ("CN-WFOE", "SG-PTE", "90500.00", "Shenzhen buffer stock"),
        ("MX-SA", "BR-LTDA", "22100.00", "LATAM quality audit"),
        ("BR-LTDA", "US-MFG", "38900.00", "Port of Santos demurrage"),
        ("AU-PTY", "CN-WFOE", "15600.00", "ANZ returns processing"),
        ("US-HOLD", "SG-PTE", "120000.00", "Regional guarantee fee"),
        ("US-HOLD", "IN-PVT", "64000.00", "IP licence India"),
        ("US-HOLD", "MX-SA", "71000.00", "IP licence Mexico"),
        ("US-HOLD", "BR-LTDA", "54000.00", "IP licence Brazil"),
        ("US-HOLD", "AU-PTY", "49000.00", "IP licence Australia"),
        ("US-HOLD", "CN-WFOE", "130000.00", "IP licence China"),
        ("US-MFG", "IN-PVT", "41000.00", "Casting drawings"),
        ("DE-NEX", "UK-LTD", "22800.00", "UK field service"),
        ("DE-NEX", "SG-PTE", "33600.00", "ASEAN application engineering"),
        ("UK-LTD", "IN-PVT", "14700.00", "GST documentation support"),
        ("FR-SAS", "MX-SA", "9800.00", "Spanish HMI skins"),
        ("NL-BV", "US-HOLD", "180000.00", "Cash-pool interest"),
        ("SG-PTE", "DE-NEX", "26500.00", "EU-APAC logistics overlay"),
        ("IN-PVT", "DE-NEX", "19800.00", "PLC software escrow"),
        ("CN-WFOE", "NL-BV", "47200.00", "Rotterdam bonded warehouse"),
        ("MX-SA", "US-HOLD", "31500.00", "IMMEX compliance"),
        ("BR-LTDA", "FR-SAS", "11200.00", "Airbus supplier onboarding"),
        ("AU-PTY", "UK-LTD", "13400.00", "Mining OEM support"),
    ]
    for i, (a, b, amt, note) in enumerate(extras, start=2600):
        day = date(2024, 1, 4) + timedelta(days=i % 26)
        pair(a, b, amt, f"IC-{i}", day, day, note, f"XR-{i}")

    # In-transit: A books 31 Jan, B books 3 Feb (inside 5-day window)
    pair("US-HOLD", "DE-NEX", "42000.00", "IC-2490", date(2024, 1, 31), date(2024, 2, 3), "In-transit timing — month-end goods", "IT-DE-31")

    # Material break: AR 1,850,000 vs AP 1,835,000 (break 15,000 > 1,000)
    pair("SG-PTE", "US-MFG", "1850000.00", "IC-2491", date(2024, 1, 14), date(2024, 1, 14), "APAC finished goods — disputed freight", "FG-US-91")

    # M:N netting: three AR invoices vs one AP payment on counterpart
    pair("NL-BV", "FR-SAS", "40000.00", "IC-2501", date(2024, 1, 8), date(2024, 1, 8), "Netting basket 1 of 3", "NT-01")
    pair("NL-BV", "FR-SAS", "35000.00", "IC-2502", date(2024, 1, 9), date(2024, 1, 9), "Netting basket 2 of 3", "NT-02")
    pair("NL-BV", "FR-SAS", "25000.00", "IC-2503", date(2024, 1, 10), date(2024, 1, 10), "Netting basket 3 of 3", "NT-03")

    by_entity: dict[str, list[list[object]]] = {e[0]: [] for e in entities}
    seq = 0
    for a, b, amt, doc, d1, d2, note, ar_inv, ap_inv in pairs:
        seq += 1
        a_rec = next(e for e in entities if e[0] == a)
        b_rec = next(e for e in entities if e[0] == b)
        ar_amt = amt
        ap_amt = amt
        if doc == "IC-2491":
            ap_amt = money("1835000.00")
        by_entity[a].append(
            [
                a,
                a_rec[1],
                name_for(a, a),
                b,
                name_for(b, a),
                "AR",
                f"{doc}-AR",
                d1,
                d1,
                ar_amt,
                "USD",
                ar_inv,
                note,
                f"{a_rec[4]}-{80000 + seq}",
                b,
                "129100 IC AR",
            ]
        )
        by_entity[b].append(
            [
                b,
                b_rec[1],
                name_for(b, b),
                a,
                name_for(a, b),
                "AP",
                f"{doc}-AP",
                d2,
                d2,
                ap_amt,
                "USD",
                ap_inv,
                note,
                f"{b_rec[4]}-{90000 + seq}",
                a,
                "229100 IC AP",
            ]
        )

    # Netting payment on FR books as one AP settlement covering NT-01..03
    fr = next(e for e in entities if e[0] == "FR-SAS")
    by_entity["FR-SAS"].append(
        [
            "FR-SAS",
            fr[1],
            name_for("FR-SAS", "FR-SAS"),
            "NL-BV",
            name_for("NL-BV", "FR-SAS"),
            "AP",
            "IC-NET-MAY-AP",
            date(2024, 1, 31),
            date(2024, 1, 31),
            money("100000.00"),
            "USD",
            "NT-01/NT-02/NT-03",
            "Single netting payment settling three NL invoices",
            "S/4-NET-01",
            "NL-BV",
            "229100 IC AP",
        ]
    )

    ledgers = out / "ledgers"
    ledgers.mkdir(exist_ok=True)
    for code, legal, *_rest in entities:
        write_xlsx(
            ledgers / f"IC_Ledger_{code}.xlsx",
            code[:31],
            headers,
            by_entity[code],
            money_idx=[9],
            date_idx=[7, 8],
        )

    master_rows = [[e[0], e[1], e[2], e[3], e[4]] for e in entities]
    write_xlsx(
        out / "IC_Entity_Master.xlsx",
        "Entity_Master",
        ["entity_code", "legal_name", "alias_seen_in_erps", "country", "local_erp"],
        master_rows,
        money_idx=[],
        date_idx=[],
    )

    write_policy_pdf(
        out / "Intercompany_Governance_Policy.pdf",
        "Intercompany Governance Policy",
        "Document IC-GOV-12  ·  12 legal entities  ·  USD reporting (no FX in v1)",
        [
            (
                "1. Directional matching",
                [
                    "Entity A’s receivable from Entity B may only match Entity B’s payable to Entity A. Do not net opposite corridors in the same matcher node (A→B is distinct from B→A).",
                ],
            ),
            (
                "2. Entity master (normalize local ERP names)",
                [
                    "US-HOLD Helios Holdings Inc | US-MFG Helios Manufacturing USA | DE-NEX Nexus Germany GmbH (alias DE-Nexus Corp) | UK-LTD Helios UK Ltd | FR-SAS Helios France SAS | NL-BV Helios Netherlands BV | SG-PTE Helios Singapore Pte Ltd | IN-PVT Helios India Pvt Ltd | CN-WFOE Helios China Co Ltd | MX-SA Helios Mexico SA de CV | BR-LTDA Helios Brasil Ltda | AU-PTY Helios Australia Pty Ltd.",
                ],
            ),
            (
                "3. In-transit window",
                [
                    "Allow a <b>5 calendar-day</b> in-transit window near month-end. Example: US-HOLD books IC-2490 on 31 January; DE-NEX books 3 February. Tag as In-Transit Timing Break, not a hard mismatch.",
                ],
            ),
            (
                "4. Netting and materiality",
                [
                    "When several invoices are settled with one payment, chain an allocation (M:N) matcher after the directional match. Example: NL-BV invoices NT-01 / NT-02 / NT-03 versus FR-SAS payment IC-NET-MAY-AP (USD 100,000).",
                    "Net AR minus AP by entity pair. Highlight material net breaks over <b>USD 1,000</b>. Known dispute: IC-2491 SG-PTE AR USD 1,850,000 versus US-MFG AP USD 1,835,000 (USD 15,000 freight).",
                ],
            ),
            (
                "5. Deliverables",
                [
                    "Treasury multi-tab Excel matrix and an executive PDF briefing for the Corporate Controller.",
                ],
            ),
        ],
    )


def journey_invoice_payments() -> None:
    """AR payment confirmation: bank statement vs customer invoices."""
    out = ROOT / "journey-04-invoice-payment-confirmation"
    out.mkdir(parents=True, exist_ok=True)

    inv_headers = [
        "Invoice Number",
        "Invoice Date",
        "Due Date",
        "Customer Name",
        "Customer ID",
        "Invoice Amount",
        "Tax Amount",
        "Invoice Total",
        "Currency",
        "Payment Terms",
        "Sales Rep",
        "Region",
        "Status",
    ]
    bank_headers = [
        "Statement Date",
        "Value Date",
        "Bank Txn Id",
        "Txn Type",
        "Debit Credit",
        "Amount",
        "Reference Number",
        "Payer Name",
        "Description",
        "Remittance Memo",
        "Branch",
        "Currency",
    ]

    invoices = [
        [
            "INV-24-6001",
            date(2024, 5, 2),
            date(2024, 6, 1),
            "Helios Stores NA",
            "C-1001",
            money("128440.15"),
            money("0.00"),
            money("128440.15"),
            "USD",
            "Net 30",
            "R. Alvarez",
            "Midwest",
            "open",
        ],
        [
            "INV-24-6002",
            date(2024, 5, 6),
            date(2024, 6, 5),
            "Northwind Wholesale Inc",
            "C-1044",
            money("48500.00"),
            money("0.00"),
            money("48500.00"),
            "USD",
            "Net 30",
            "T. Berg",
            "East",
            "open",
        ],
        [
            "INV-24-6003",
            date(2024, 5, 8),
            date(2024, 6, 7),
            "Acme Distributors LLC",
            "C-1102",
            money("45200.00"),
            money("0.00"),
            money("45200.00"),
            "USD",
            "Net 30",
            "H. Cole",
            "Midwest",
            "open",
        ],
        [
            "INV-24-6004",
            date(2024, 5, 10),
            date(2024, 6, 9),
            "Contoso Manufacturing",
            "C-1188",
            money("88400.00"),
            money("0.00"),
            money("88400.00"),
            "USD",
            "Net 45",
            "R. Alvarez",
            "West",
            "open",
        ],
        [
            "INV-24-6005",
            date(2024, 5, 12),
            date(2024, 6, 11),
            "Fastenal Company",
            "C-1210",
            money("12000.00"),
            money("0.00"),
            money("12000.00"),
            "USD",
            "Net 15",
            "T. Berg",
            "Midwest",
            "open",
        ],
        [
            "INV-24-6006",
            date(2024, 5, 14),
            date(2024, 6, 13),
            "Dealer AR Midwest",
            "C-1308",
            money("32000.00"),
            money("0.00"),
            money("32000.00"),
            "USD",
            "Net 30",
            "H. Cole",
            "Midwest",
            "open",
        ],
        [
            "INV-24-6007",
            date(2024, 4, 20),
            date(2024, 5, 20),
            "Old Dominion Freight",
            "C-1412",
            money("22100.00"),
            money("0.00"),
            money("22100.00"),
            "USD",
            "Net 30",
            "R. Alvarez",
            "East",
            "past_due",
        ],
        [
            "INV-24-6008",
            date(2024, 5, 18),
            date(2024, 6, 17),
            "ABM Industry",
            "C-1550",
            money("6400.00"),
            money("0.00"),
            money("6400.00"),
            "USD",
            "Net 30",
            "T. Berg",
            "Midwest",
            "open",
        ],
    ]

    bank_rows: list[list[object]] = []
    bid = 81000

    def add_bank(d, amt, ref, payer, desc, memo, ttype="deposit", dc="CR"):
        nonlocal bid
        bid += 1
        bank_rows.append(
            [
                d,
                d,
                f"BNK-{bid}",
                ttype,
                dc,
                money(amt),
                ref,
                payer,
                desc,
                memo,
                "Cincinnati OH",
                "USD",
            ]
        )

    add_bank(date(2024, 6, 3), "128440.15", "INV-24-6001", "Helios Stores NA", "Card settlement — Helios Retail POS", "INV-24-6001 exact")
    add_bank(date(2024, 6, 4), "48500.00", "INV-24-6002", "NORTHWIND WHSLE CO", "Incoming wire — customer remittance", "INV-24-6002 name variant")
    add_bank(date(2024, 6, 5), "45190.00", "INV-24-6003", "Acme Distributors LLC", "ACH customer payment", "INV-24-6003 short $10")
    add_bank(date(2024, 6, 6), "79560.00", "INV-24-6004", "Contoso Manufacturing", "ACH customer payment", "INV-24-6004 short $8,840")
    add_bank(date(2024, 6, 7), "15000.00", "INV-24-6005", "Fastenal Company", "Vendor rebate / overpay", "INV-24-6005 overpay $3,000")
    add_bank(date(2024, 6, 8), "20000.00", "INV-24-6006-A", "Dealer AR Midwest", "Lockbox — partial 1 of 2", "INV-24-6006 split 20000")
    add_bank(date(2024, 6, 10), "12000.00", "INV-24-6006-B", "Dealer AR Midwest", "Lockbox — partial 2 of 2", "INV-24-6006 split 12000")
    add_bank(date(2024, 6, 11), "6400.00", "INV-24-6008", "ABM Industry", "Facilities remittance", "INV-24-6008 first pay")
    add_bank(date(2024, 6, 14), "6400.00", "INV-24-6008-DUP", "ABM Industry", "Duplicate remittance", "INV-24-6008 duplicate")
    add_bank(date(2024, 6, 12), "18000.00", "WIRE-NEX-8801", "Grainger Inc", "Unrelated incoming wire — no invoice", "operating noise")

    noise = [
        (date(2024, 6, 3), "882140.00", "ACH-PAY-0603", "ADP Wage Pay", "Bi-weekly payroll funding", "payroll", "withdrawal", "DR"),
        (date(2024, 6, 4), "44120.66", "ACH-UTIL-88", "Midwest Power", "Electric — plants OH/IN/TX", "utilities", "withdrawal", "DR"),
        (date(2024, 6, 5), "12880.00", "ACH-LEASE-04", "Prologis REIT", "Distribution center lease", "lease", "withdrawal", "DR"),
        (date(2024, 6, 6), "6720.45", "FEE-ANA-JUN", "First National Treasury", "Account analysis fee", "bank fee", "withdrawal", "DR"),
        (date(2024, 6, 7), "9104.22", "ACH-INS-Q2", "Chubb North America", "Property insurance Q2", "insurance", "withdrawal", "DR"),
        (date(2024, 6, 10), "33450.00", "WIRE-TAX-941", "US Treasury", "Federal 941 remittance", "tax", "withdrawal", "DR"),
        (date(2024, 6, 11), "15600.80", "ACH-BENE-HSA", "Fidelity Benefits", "HSA employer funding", "benefits", "withdrawal", "DR"),
        (date(2024, 6, 12), "3122.18", "ACH-COURIER", "FedEx Corporate", "Overnight freight", "courier", "withdrawal", "DR"),
        (date(2024, 6, 13), "11880.00", "ACH-CLOUD", "SAP America", "ERP cloud subscription", "saas", "withdrawal", "DR"),
        (date(2024, 6, 14), "775.40", "FEE-WIRE-OUT", "First National Treasury", "Wire origination fee", "wire fee", "withdrawal", "DR"),
        (date(2024, 6, 15), "250000.00", "SWEEP-MMF", "Federated Hermes", "Overnight money-market sweep", "sweep", "withdrawal", "DR"),
        (date(2024, 6, 17), "4550.00", "WIRE-REBATE", "Industrial Supply Rebate Co", "Vendor rebate — operating", "rebate noise", "deposit", "CR"),
        (date(2024, 6, 18), "990.00", "NSF-2291", "Returned Item NSF", "Returned item NSF", "nsf", "withdrawal", "DR"),
        (date(2024, 6, 19), "6400.00", "ACH-JANITOR", "ABM Industry", "Facilities services outflow", "ap payment", "withdrawal", "DR"),
    ]
    for d, amt, ref, payer, desc, memo, ttype, dc in noise:
        add_bank(d, amt, ref, payer, desc, memo, ttype, dc)

    write_xlsx(out / "Customer_Invoices.xlsx", "Invoices", inv_headers, invoices, money_idx=[5, 6, 7], date_idx=[1, 2])
    write_xlsx(out / "Bank_Statement_June.xlsx", "Bank_June", bank_headers, bank_rows, money_idx=[5], date_idx=[0, 1])
    write_policy_pdf(
        out / "AR_Collections_Policy.pdf",
        "AR Collections & Cash Application Policy",
        "Document AR-SOP-2024.06  ·  Effective 1 June 2024  ·  Helios Industrial",
        [
            (
                "1. Purpose",
                [
                    "Confirm which customer invoices were paid from the bank statement. The invoice register is the source of truth for what is owed. The bank statement contains many unrelated operating items (payroll, utilities, fees, sweeps) that must not be treated as customer receipts.",
                ],
            ),
            (
                "2. Match rule (Matcher — names, not amounts)",
                [
                    "Pair invoice rows to bank receipts on <b>customer name</b> / <b>payer name</b>. Legal names may differ from lockbox names (for example Northwind Wholesale Inc versus NORTHWIND WHSLE CO). Use semantic normalization. Do not fail a match solely on spelling.",
                    "Do not use amount as a match key. Short payments and overpayments must still pair so Finance can see the variance.",
                ],
            ),
            (
                "3. Amount tolerance (Math Engine)",
                [
                    "After a name match, compare invoice <b>total_amount</b> to the paid bank <b>amount</b>. Allow the lesser of <b>1% of the invoice total</b> or <b>USD 50.00</b>. Formula: abs(actual − expected) &gt; min(0.01 × expected, 50.00) flags the row.",
                    "Example in-tolerance: invoice USD 45,200.00 vs receipt USD 45,190.00 (USD 10.00 under, under the USD 50 cap). Example out-of-tolerance: invoice USD 88,400.00 vs receipt USD 79,560.00.",
                ],
            ),
            (
                "4. Split payments and duplicates",
                [
                    "One invoice may be settled by several deposits. Match them as a group and emit leftovers on residuals. One wire covering two invoices is the opposite allocation. A second payment of the same invoice amount is a duplicate and must be flagged.",
                ],
            ),
            (
                "5. Aging and unpaid invoices",
                [
                    "Invoices with no bank receipt that are past due more than <b>30 calendar days</b> escalate for collections. Cite this SOP in the decision explanation. INV-24-6007 Old Dominion Freight USD 22,100.00 is the in-scope unpaid example.",
                ],
            ),
            (
                "6. Output",
                [
                    "Excel workbook with a sheet per outcome: matched (paid), residuals (partial / split leftovers), exceptions (unpaid, duplicate, out-of-tolerance). Download only — no Slack.",
                ],
            ),
        ],
    )


def main() -> None:
    journey_invoice()
    journey_bank()
    journey_intercompany()
    journey_invoice_payments()
    print("Wrote", ROOT)


if __name__ == "__main__":
    main()
