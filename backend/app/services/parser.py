from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from pypdf import PdfReader

SUPPORTED = {".csv", ".xlsx", ".pdf", ".txt", ".md"}
ColumnType = Literal["string", "integer", "decimal", "date", "boolean"]

_DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%Y/%m/%d",
    "%d %b %Y",
    "%b %d, %Y",
)


class ParseError(ValueError):
    """File cannot be read as a supported table or document."""


@dataclass
class DetectedColumn:
    name: str
    type: ColumnType
    samples: list[Any]


@dataclass
class ParsedTable:
    columns: list[DetectedColumn]
    rows: list[dict[str, Any]]
    sheet: str | None = None
    header_row: int = 1


@dataclass
class ParseResult:
    path: str
    fmt: str
    text: str
    tables: list[ParsedTable] = field(default_factory=list)


def parse_file(
    path: Path,
    *,
    sheet: str | None = None,
    header_row: int | None = None,
) -> ParseResult:
    path = Path(path)
    if not path.exists():
        raise ParseError(f"file not found: {path}")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED:
        raise ParseError(f"unsupported format {suffix}; expected .xlsx, .csv, .pdf, .txt, or .md")
    try:
        if suffix == ".csv":
            return _parse_csv(path, header_row)
        if suffix == ".xlsx":
            return _parse_xlsx(path, sheet, header_row)
        if suffix in {".txt", ".md"}:
            return _parse_text(path)
        return _parse_pdf(path, header_row)
    except ParseError:
        raise
    except Exception as exc:
        raise ParseError(f"malformed {suffix} file: {exc}") from exc


def _parse_text(path: Path) -> ParseResult:
    text = path.read_text(encoding="utf-8", errors="replace")
    return ParseResult(path=str(path), fmt=path.suffix.lstrip("."), text=text)


def detect_schema(rows: list[dict[str, Any]], headers: list[str]) -> list[DetectedColumn]:
    columns: list[DetectedColumn] = []
    for header in headers:
        values = [row.get(header) for row in rows]
        samples = [v for v in values if v not in (None, "")][:5]
        columns.append(
            DetectedColumn(name=header, type=_infer_type(values), samples=samples)
        )
    return columns


def apply_overrides(
    columns: list[DetectedColumn],
    rows: list[dict[str, Any]],
    overrides: dict[str, dict[str, Any]] | None,
) -> tuple[list[DetectedColumn], list[dict[str, Any]]]:
    if not overrides:
        return columns, rows
    rename = {
        old: patch["name"]
        for old, patch in overrides.items()
        if "name" in patch and patch["name"] != old
    }
    new_cols: list[DetectedColumn] = []
    for col in columns:
        patch = overrides.get(col.name, {})
        name = patch.get("name", col.name)
        ctype = patch.get("type", col.type)
        new_cols.append(DetectedColumn(name=name, type=ctype, samples=col.samples))
    new_rows = []
    for row in rows:
        item = {}
        for key, value in row.items():
            item[rename.get(key, key)] = value
        new_rows.append(item)
    return new_cols, new_rows


def columns_as_dicts(columns: list[DetectedColumn]) -> list[dict[str, Any]]:
    return [
        {"name": c.name, "type": c.type, "samples": c.samples}
        for c in columns
    ]


def _parse_csv(path: Path, header_row: int | None) -> ParseResult:
    raw = path.read_bytes()
    if not raw.strip():
        raise ParseError("csv file is empty")
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    if "\x00" in text[:200] and "," not in text[:200]:
        raise ParseError("csv file looks binary")
    reader = csv.reader(io.StringIO(text))
    matrix = [row for row in reader]
    table = _table_from_matrix(matrix, header_row, sheet=None)
    return ParseResult(path=str(path), fmt="csv", text=text, tables=[table])


def _parse_xlsx(path: Path, sheet: str | None, header_row: int | None) -> ParseResult:
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise ParseError(f"malformed xlsx: {exc}") from exc
    if sheet:
        if sheet not in wb.sheetnames:
            raise ParseError(f"sheet {sheet!r} not found")
        ws = wb[sheet]
    else:
        ws = wb.active
        sheet = ws.title
    matrix: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        matrix.append(list(row))
    table = _table_from_matrix(matrix, header_row, sheet=sheet)
    text_lines = [",".join("" if c is None else str(c) for c in row) for row in matrix]
    return ParseResult(
        path=str(path),
        fmt="xlsx",
        text="\n".join(text_lines),
        tables=[table],
    )


def _parse_pdf(path: Path, header_row: int | None) -> ParseResult:
    try:
        reader = PdfReader(str(path))
    except Exception as exc:
        raise ParseError(f"malformed pdf: {exc}") from exc
    if getattr(reader, "is_encrypted", False):
        raise ParseError("encrypted pdf is not supported")
    pages = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    text = "\n".join(pages).strip()
    if not text:
        raise ParseError("pdf contained no extractable text")
    matrix = _matrix_from_text(text)
    tables = []
    if matrix:
        try:
            tables = [_table_from_matrix(matrix, header_row, sheet=None)]
        except ParseError:
            tables = []
    return ParseResult(path=str(path), fmt="pdf", text=text, tables=tables)


def _matrix_from_text(text: str) -> list[list[str]]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return []
    if any("," in ln for ln in lines[:8]):
        return [next(csv.reader([ln])) for ln in lines]
    if any("\t" in ln for ln in lines[:8]):
        return [ln.split("\t") for ln in lines]
    split = [re.split(r"\s{2,}", ln) for ln in lines]
    if split and max(len(r) for r in split) >= 2:
        return split
    return [[ln] for ln in lines]


def _table_from_matrix(
    matrix: list[list[Any]],
    header_row: int | None,
    sheet: str | None,
) -> ParsedTable:
    cleaned: list[list[Any]] = []
    for row in matrix:
        values = list(row)
        if all(v is None or str(v).strip() == "" for v in values):
            if cleaned:
                continue
            continue
        cleaned.append(values)
    if not cleaned:
        raise ParseError("no header or data rows found")
    idx = (header_row - 1) if header_row else 0
    if idx < 0 or idx >= len(cleaned):
        raise ParseError("header_row is out of range")
    headers = [_unique_header(c, i) for i, c in enumerate(cleaned[idx])]
    if not any(h.strip() for h in headers):
        raise ParseError("header row is empty")
    data_rows = cleaned[idx + 1 :]
    rows: list[dict[str, Any]] = []
    for raw in data_rows:
        if all(v is None or str(v).strip() == "" for v in raw):
            continue
        item: dict[str, Any] = {}
        for i, header in enumerate(headers):
            value = raw[i] if i < len(raw) else None
            item[header] = _normalize_cell(value)
        rows.append(item)
    columns = detect_schema(rows, headers)
    return ParsedTable(
        columns=columns,
        rows=rows,
        sheet=sheet,
        header_row=idx + 1,
    )


def _unique_header(cell: Any, index: int) -> str:
    name = "" if cell is None else str(cell).strip()
    return name or f"column_{index + 1}"


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, float):
        return str(Decimal(str(value)))
    text = str(value).strip()
    return text if text != "" else None


def _infer_type(values: list[Any]) -> ColumnType:
    nonempty = [v for v in values if v not in (None, "")]
    if not nonempty:
        return "string"
    if all(_is_bool(v) for v in nonempty):
        return "boolean"
    if all(_is_int(v) for v in nonempty):
        return "integer"
    if all(_is_decimal(v) for v in nonempty):
        return "decimal"
    if all(_is_date(v) for v in nonempty):
        return "date"
    return "string"


def _is_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return True
    return str(value).strip().lower() in {"true", "false", "yes", "no", "y", "n"}


def _is_int(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    text = str(value).strip().replace(",", "")
    return bool(re.fullmatch(r"[+-]?\d+", text))


def _is_decimal(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float, Decimal)):
        return True
    text = str(value).strip().replace(",", "").replace("$", "")
    try:
        Decimal(text)
        return True
    except (InvalidOperation, ValueError):
        return False


def _is_date(value: Any) -> bool:
    if isinstance(value, (date, datetime)):
        return True
    text = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", text))
