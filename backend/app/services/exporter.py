from __future__ import annotations

import json
import re
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfReader
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from app.models.envelope import Envelope


@dataclass(frozen=True)
class Theme:
    id: str
    label: str
    header: str
    accent: str
    text: str
    paper: str
    good: str
    bad: str
    muted: str


THEMES: dict[str, Theme] = {
    "executive_classic": Theme(
        id="executive_classic",
        label="Executive Classic",
        header="1B3A4B",
        accent="C4A35A",
        text="1A1A1A",
        paper="F7F4EF",
        good="2E7D4F",
        bad="A33B3B",
        muted="6B7280",
    ),
    "modern_slate": Theme(
        id="modern_slate",
        label="Modern Slate",
        header="1F2937",
        accent="3B82F6",
        text="111827",
        paper="F3F4F6",
        good="059669",
        bad="DC2626",
        muted="6B7280",
    ),
    "audit_clean": Theme(
        id="audit_clean",
        label="Audit Clean",
        header="14532D",
        accent="4D7C0F",
        text="052E16",
        paper="FFFFFF",
        good="166534",
        bad="9F1239",
        muted="57534E",
    ),
}


@dataclass
class Stream:
    name: str
    port: str
    kind: str
    rows: list[dict[str, Any]]


def resolve_theme(name: str | None) -> Theme:
    key = (name or "executive_classic").lower().replace(" ", "_").replace("-", "_")
    aliases = {
        "executive": "executive_classic",
        "classic": "executive_classic",
        "slate": "modern_slate",
        "modern": "modern_slate",
        "audit": "audit_clean",
        "clean": "audit_clean",
    }
    resolved = aliases.get(key, key)
    return THEMES.get(resolved, THEMES["executive_classic"])


def collect_streams(inputs: list[Envelope], config: dict[str, Any]) -> list[Stream]:
    labels = dict(config.get("tabs") or {})
    streams: list[Stream] = []
    used: set[str] = set()
    for index, env in enumerate(inputs):
        payload = env.payload or {}
        if payload.get("kind") == "knowledge":
            continue
        rows = payload.get("rows")
        if not isinstance(rows, list):
            continue
        key = env.port if env.port != "default" else payload.get("kind") or f"stream_{index + 1}"
        name = labels.get(env.port) or labels.get(str(index)) or _sheet_name(str(key), used)
        used.add(name.lower())
        streams.append(
            Stream(
                name=name,
                port=env.port,
                kind=str(payload.get("kind") or "table"),
                rows=[flatten_row(row) if isinstance(row, dict) else {"value": row} for row in rows],
            )
        )
    if not streams:
        streams.append(Stream(name="Data", port="default", kind="empty", rows=[]))
    return streams


def flatten_row(row: dict[str, Any], prefix: str = "") -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in row.items():
        name = f"{prefix}{key}" if prefix else str(key)
        if isinstance(value, dict):
            out.update(flatten_row(value, f"{name}."))
        elif isinstance(value, list):
            out[name] = json.dumps(value, default=str)
        elif isinstance(value, Decimal):
            out[name] = float(value)
        else:
            out[name] = value
    return out


def summarize(streams: list[Stream]) -> dict[str, Any]:
    total = sum(len(s.rows) for s in streams)
    by_port = {s.port: len(s.rows) for s in streams}
    statuses: dict[str, int] = {}
    for stream in streams:
        for row in stream.rows:
            label = str(row.get("verdict") or row.get("status") or stream.port or "row")
            statuses[label] = statuses.get(label, 0) + 1
    return {
        "streams": len(streams),
        "rows": total,
        "by_port": by_port,
        "statuses": statuses,
    }


def write_workbook(streams: list[Stream], *, theme: Theme, title: str) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Executive Summary"
    kpis = summarize(streams)
    summary.append(["Report", title])
    summary.append(["Theme", theme.label])
    summary.append(["Streams", kpis["streams"]])
    summary.append(["Rows", kpis["rows"]])
    summary.append([])
    summary.append(["KPI", "Value"])
    _header_row(summary, 6, theme, 2)
    for i, (key, value) in enumerate(kpis["by_port"].items(), start=7):
        summary.append([f"{key} rows", value])
        _zebra(summary, i, theme)
    summary.freeze_panes = "A2"
    _autosize(summary)

    used = {ws.title.lower() for ws in wb.worksheets}
    for stream in streams:
        ws = wb.create_sheet(_sheet_name(stream.name, used))
        used.add(ws.title.lower())
        _write_table(ws, stream, theme)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_pdf(
    streams: list[Stream],
    *,
    theme: Theme,
    title: str,
    charts: dict[str, bool] | None = None,
) -> bytes:
    charts = {"donut": True, "variance": True, "trend": True, **(charts or {})}
    kpis = summarize(streams)
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter
    c.setTitle(title)
    _pdf_header(c, width, height, title, theme)
    y = height - 1.15 * inch
    y = _kpi_cards(c, y, kpis, theme)
    chart_y = y - 10
    if charts.get("donut"):
        _donut(c, 0.75 * inch, chart_y, kpis["statuses"] or {"rows": max(kpis["rows"], 1)}, theme)
    if charts.get("variance"):
        _bars(c, 4.1 * inch, chart_y, streams, theme)
    y = chart_y - 20
    if charts.get("trend"):
        y = _trend(c, 0.75 * inch, y, streams, theme)
    _pdf_table(c, 0.6 * inch, min(y, 250), streams, theme)
    _signoff(c, width, theme)
    c.showPage()
    c.save()
    return buf.getvalue()


def inspect_workbook(data: bytes) -> dict[str, Any]:
    wb = load_workbook(BytesIO(data))
    sheets = []
    fills: list[str] = []
    for ws in wb.worksheets:
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.value)
                fill = getattr(cell.fill, "fgColor", None)
                rgb = getattr(fill, "rgb", None) if fill is not None else None
                if rgb and str(rgb) not in {"00000000", "0"}:
                    fills.append(str(rgb)[-6:].upper())
        sheets.append(
            {
                "name": ws.title,
                "freeze": str(ws.freeze_panes),
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "formulas": formulas,
            }
        )
    return {"sheets": sheets, "header_fills": sorted(set(fills))}


def inspect_pdf(data: bytes) -> dict[str, Any]:
    reader = PdfReader(BytesIO(data))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return {"pages": len(reader.pages), "text": text}


def _write_table(ws: Worksheet, stream: Stream, theme: Theme) -> None:
    rows = stream.rows
    columns = _columns(rows)
    ws.append(columns or ["(empty)"])
    _header_row(ws, 1, theme, max(len(columns), 1))
    ws.freeze_panes = "A2"
    numeric: set[int] = set()
    status_col = None
    for r_i, row in enumerate(rows, start=2):
        ws.append([_cell(row.get(col)) for col in columns])
        for c_i, col in enumerate(columns, start=1):
            if _is_number(row.get(col)):
                numeric.add(c_i)
                ws.cell(r_i, c_i).number_format = "#,##0.00"
            if status_col is None and col.lower() in {"status", "verdict", "flag"}:
                status_col = c_i
        _zebra(ws, r_i, theme)
    data_end = 1 + len(rows)
    if numeric and rows:
        formula_row = data_end + 1
        ws.cell(formula_row, 1, "Total")
        for col_i in numeric:
            letter = get_column_letter(col_i)
            ws.cell(formula_row, col_i, f"=SUM({letter}2:{letter}{data_end})")
            ws.cell(formula_row, col_i).font = Font(bold=True, color=theme.header)
        ws.cell(formula_row, 1).font = Font(bold=True, color=theme.header)
    if status_col and rows:
        letter = get_column_letter(status_col)
        rng = f"{letter}2:{letter}{data_end}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=['"matched"'],
                fill=PatternFill("solid", fgColor=theme.good),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'OR({letter}2="approved",{letter}2="Approved")'],
                fill=PatternFill("solid", fgColor=theme.good),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f'OR({letter}2="exceptions",{letter}2="flagged",{letter}2="escalated",{letter}2="unmatched")'
                ],
                fill=PatternFill("solid", fgColor=theme.bad),
            ),
        )
    _autosize(ws)
    ws.sheet_properties.tabColor = theme.accent


def _columns(rows: list[dict[str, Any]]) -> list[str]:
    seen: list[str] = []
    for row in rows:
        for key in row:
            if key not in seen:
                seen.append(key)
    return seen


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _is_number(value: Any) -> bool:
    if isinstance(value, bool) or value in (None, ""):
        return False
    if isinstance(value, (int, float, Decimal)):
        return True
    try:
        float(str(value).replace(",", "").replace("$", ""))
        return True
    except (TypeError, ValueError):
        return False


def _header_row(ws: Worksheet, row: int, theme: Theme, width: int) -> None:
    fill = PatternFill("solid", fgColor=theme.header)
    font = Font(bold=True, color="FFFFFF", name="Calibri")
    for col in range(1, width + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _zebra(ws: Worksheet, row: int, theme: Theme) -> None:
    if row % 2 == 0:
        return
    fill = PatternFill("solid", fgColor=theme.paper)
    for cell in ws[row]:
        rgb = str(getattr(getattr(cell.fill, "fgColor", None), "rgb", "") or "")
        if rgb in {"00000000", "0", ""}:
            cell.fill = fill


def _autosize(ws: Worksheet) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = 12
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = min(max(width, len(value) + 2), 42)
        ws.column_dimensions[letter].width = width
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions


def _sheet_name(raw: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", raw).strip() or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    n = 1
    while cleaned.lower() in used:
        n += 1
        suffix = f"_{n}"
        cleaned = base[: 31 - len(suffix)] + suffix
    return cleaned


def _hex(color: str) -> tuple[float, float, float]:
    color = color.lstrip("#")
    return (int(color[0:2], 16) / 255, int(color[2:4], 16) / 255, int(color[4:6], 16) / 255)


def _pdf_header(c: canvas.Canvas, width: float, height: float, title: str, theme: Theme) -> None:
    c.setFillColorRGB(*_hex(theme.header))
    c.rect(0, height - 0.7 * inch, width, 0.7 * inch, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Times-Bold", 16)
    c.drawString(0.6 * inch, height - 0.45 * inch, title)
    c.setFont("Helvetica", 9)
    c.drawRightString(width - 0.5 * inch, height - 0.42 * inch, theme.label)


def _kpi_cards(c: canvas.Canvas, y: float, kpis: dict[str, Any], theme: Theme) -> float:
    items = [("Streams", kpis["streams"]), ("Rows", kpis["rows"])]
    items.extend((str(k), v) for k, v in list(kpis["statuses"].items())[:2])
    card_w = 1.7 * inch
    x = 0.55 * inch
    for label, value in items[:4]:
        c.setFillColorRGB(*_hex(theme.paper))
        c.setStrokeColorRGB(*_hex(theme.accent))
        c.roundRect(x, y - 0.7 * inch, card_w, 0.75 * inch, 6, fill=1, stroke=1)
        c.setFillColorRGB(*_hex(theme.muted))
        c.setFont("Helvetica", 8)
        c.drawString(x + 8, y - 0.18 * inch, str(label)[:18])
        c.setFillColorRGB(*_hex(theme.header))
        c.setFont("Times-Bold", 16)
        c.drawString(x + 8, y - 0.5 * inch, str(value))
        x += card_w + 0.15 * inch
    return y - 1.05 * inch


def _donut(c: canvas.Canvas, x: float, y: float, counts: dict[str, int], theme: Theme) -> None:
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(*_hex(theme.text))
    c.drawString(x, y, "Match / status")
    total = sum(counts.values()) or 1
    colors = [theme.good, theme.accent, theme.bad, theme.muted, theme.header]
    start = 90
    cx, cy, rad = x + 50, y - 48, 32
    for i, (label, n) in enumerate(counts.items()):
        extent = 360.0 * (n / total)
        c.setFillColorRGB(*_hex(colors[i % len(colors)]))
        c.wedge(cx - rad, cy - rad, cx + rad, cy + rad, start, extent, fill=1, stroke=0)
        start += extent
        c.setFillColorRGB(*_hex(theme.text))
        c.setFont("Helvetica", 8)
        c.drawString(x + 95, y - 18 - i * 12, f"{label}: {n}")
    c.setFillColorRGB(*_hex(theme.paper))
    c.circle(cx, cy, 14, fill=1, stroke=0)


def _bars(c: canvas.Canvas, x: float, y: float, streams: list[Stream], theme: Theme) -> None:
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(*_hex(theme.text))
    c.drawString(x, y, "Variance")
    values: list[tuple[str, float]] = []
    for stream in streams:
        for row in stream.rows:
            for key, value in row.items():
                if "variance" in key.lower() and _is_number(value):
                    vendor = str(row.get("source.vendor") or row.get("vendor") or key)
                    values.append((vendor, abs(float(value))))
    values = sorted(values, key=lambda item: item[1], reverse=True)[:5]
    if not values:
        c.setFont("Helvetica", 8)
        c.setFillColorRGB(*_hex(theme.muted))
        c.drawString(x, y - 20, "No variance column")
        return
    peak = max(v for _, v in values) or 1
    for i, (label, val) in enumerate(values):
        by = y - 28 - i * 16
        c.setFillColorRGB(*_hex(theme.accent))
        c.rect(x, by, 160 * (val / peak), 10, fill=1, stroke=0)
        c.setFillColorRGB(*_hex(theme.text))
        c.setFont("Helvetica", 7)
        c.drawString(x, by + 12, f"{label[:18]} {val:.2f}")


def _trend(c: canvas.Canvas, x: float, y: float, streams: list[Stream], theme: Theme) -> float:
    points: list[tuple[int, float]] = []
    i = 0
    for stream in streams:
        for row in stream.rows:
            amount = None
            for key, value in row.items():
                if key.lower() in {"amount", "running_balance", "balance"} and _is_number(value):
                    amount = float(value)
                    break
            if amount is not None:
                points.append((i, amount))
                i += 1
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(*_hex(theme.text))
    c.drawString(x, y, "Balance / trend")
    if len(points) < 2:
        c.setFont("Helvetica", 8)
        c.setFillColorRGB(*_hex(theme.muted))
        c.drawString(x, y - 16, "Not enough points")
        return y - 40
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    min_y, max_y = min(ys), max(ys)
    span = (max_y - min_y) or 1
    c.setStrokeColorRGB(*_hex(theme.accent))
    c.setLineWidth(1.5)
    path = c.beginPath()
    for idx, (px, py) in enumerate(points):
        xx = x + (px - min(xs)) / (max(xs) - min(xs) or 1) * 220
        yy = y - 70 + (py - min_y) / span * 50
        if idx == 0:
            path.moveTo(xx, yy)
        else:
            path.lineTo(xx, yy)
    c.drawPath(path, stroke=1, fill=0)
    return y - 90


def _pdf_table(c: canvas.Canvas, x: float, y: float, streams: list[Stream], theme: Theme) -> None:
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(*_hex(theme.text))
    c.drawString(x, y, "Detail (first stream)")
    if not streams or not streams[0].rows:
        return
    cols = _columns(streams[0].rows)[:6]
    y -= 14
    c.setFont("Helvetica-Bold", 7)
    c.setFillColorRGB(*_hex(theme.header))
    c.drawString(x, y, "  ".join(col[:12] for col in cols))
    c.setFont("Helvetica", 7)
    c.setFillColorRGB(*_hex(theme.text))
    for row in streams[0].rows[:8]:
        y -= 11
        if y < 90:
            break
        c.drawString(x, y, "  ".join(str(row.get(col, ""))[:12] for col in cols))


def _signoff(c: canvas.Canvas, width: float, theme: Theme) -> None:
    y = 0.7 * inch
    c.setStrokeColorRGB(*_hex(theme.header))
    c.line(0.6 * inch, y + 36, width - 0.6 * inch, y + 36)
    c.setFillColorRGB(*_hex(theme.text))
    c.setFont("Times-Bold", 10)
    c.drawString(0.6 * inch, y + 20, "Sign-off")
    c.setFont("Helvetica", 8)
    c.drawString(0.6 * inch, y + 6, "Prepared by: ______________________")
    c.drawString(3.4 * inch, y + 6, "Reviewed by: ______________________")
    c.drawString(6.1 * inch, y + 6, "Date: ________")
